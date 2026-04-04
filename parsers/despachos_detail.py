"""
Parser del archivo de Detalle de Despachos para el sistema de conciliación.

Estructura esperada (confirmada con archivos de despachos de la gasolinera):
  - Hoja: "Reporte"
  - Headers: fila 12
  - Datos: fila 13 en adelante
  - Columnas clave:
      D (4)  = Fecha (serial Excel o datetime)
      E (5)  = Monto (colones)
      F (6)  = Volumen (litros)
      H (8)  = Combustible ("Super", "Plus 91", "Diesel", "LPG")
      I (9)  = Condicion Venta ("Credito" | "Contado")
      M (13) = Pistero (con posibles espacios y acentos)

Día laboral: 05:00 → 05:00 del día siguiente.
Los despachos entre 00:00 y 04:59 pertenecen al día laboral anterior.
"""

from __future__ import annotations

import hashlib
import re
import unicodedata
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils.datetime import from_excel

HEADER_ROW = 12
DATA_START_ROW = 13
SHEET_NAME = "Reporte"

COL_FECHA = 4
COL_MONTO = 5
COL_VOLUMEN = 6
COL_COMBUSTIBLE = 8
COL_CONDICION = 9
COL_PISTERO = 13

_WORK_DAY_START = time(5, 0)
_NON_PERSON_RE = re.compile(r"[.\d]")

# Aliases para normalización de nombres (typos conocidos en el archivo de despachos)
NAME_ALIASES: dict[str, str] = {
    "KENDALL": "KENDAL",
    "RANDALL": "RANDAL",
    "FABIAN": "FABIAN",
    "JOSE": "JOSE",
    "TOMAS": "TOMAS",
}


def normalize_name(name: str) -> str:
    if not name:
        return ""
    name = name.strip().upper()
    name = unicodedata.normalize("NFD", name).encode("ascii", "ignore").decode()
    return NAME_ALIASES.get(name, name)


def _to_datetime(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)):
        try:
            return from_excel(value)
        except Exception:
            return None
    return None


def _work_date(ts: datetime) -> date:
    cal = ts.date()
    if ts.time() < _WORK_DAY_START:
        return cal - timedelta(days=1)
    return cal


def parse_despachos(filepath: str) -> list[dict]:
    """
    Lee el archivo de Detalle de Despachos y retorna una lista de dicts.

    Cada dict tiene:
        pistero         str   nombre normalizado
        pistero_raw     str   nombre original
        work_date       date  fecha laboral (05:00→05:00)
        dispatched_at   datetime
        combustible     str
        monto           float colones
        litros          float
        es_credito      bool

    Raises:
        FileNotFoundError, ValueError
    """
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"Archivo no encontrado: {filepath}")

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError(f"No se pudo abrir el archivo: {exc}") from exc

    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"Hoja '{SHEET_NAME}' no encontrada. Hojas disponibles: {wb.sheetnames}"
        )

    ws = wb[SHEET_NAME]
    rows: list[dict] = []
    omitidos = 0

    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        if all(cell is None for cell in row):
            break
        try:
            pistero_raw = str(row[COL_PISTERO - 1] or "").strip()
            combustible = str(row[COL_COMBUSTIBLE - 1] or "").strip()
            condicion = str(row[COL_CONDICION - 1] or "").strip()
            volumen_val = row[COL_VOLUMEN - 1]
            monto_val = row[COL_MONTO - 1]
            fecha_val = row[COL_FECHA - 1]

            if not pistero_raw or _NON_PERSON_RE.search(pistero_raw):
                omitidos += 1
                continue
            if not combustible:
                omitidos += 1
                continue

            ts = _to_datetime(fecha_val)
            if ts is None:
                omitidos += 1
                continue

            litros = float(volumen_val) if volumen_val is not None else 0.0
            if litros <= 0:
                omitidos += 1
                continue

            monto = float(monto_val) if monto_val is not None else 0.0

            rows.append({
                "pistero": normalize_name(pistero_raw),
                "pistero_raw": pistero_raw,
                "work_date": _work_date(ts),
                "dispatched_at": ts,
                "combustible": combustible,
                "monto": monto,
                "litros": litros,
                "es_credito": condicion.lower() == "credito",
            })
        except Exception:
            omitidos += 1

    wb.close()
    return rows


def sha256_file(filepath: str) -> str:
    h = hashlib.sha256()
    with open(filepath, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()
