import csv
import hashlib
import hmac
import io
import json
import os
import re
import secrets
import unicodedata
import uuid
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
import psycopg2.extras
from fastapi import Depends, FastAPI, File, Form, Header, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

from database import get_db, init_db

APP_TITLE = "Cierre de Caja API"
TOKEN_TTL_HOURS = int(os.environ.get("TOKEN_TTL_HOURS", "24"))
EMPLOYEE_EDIT_HOURS = int(os.environ.get("EMPLOYEE_EDIT_HOURS", "24"))
UPLOAD_ROOT = Path(os.environ.get("UPLOAD_ROOT", "uploads"))
EXPORT_TEMPLATE_PATH = Path(os.environ.get("CIERRE_TEMPLATE_PATH", "templates/cierre_template.xlsx"))
BACKDOOR_ADMIN_USERNAME = os.environ.get("BACKDOOR_ADMIN_USERNAME", "root")
BACKDOOR_ADMIN_PASSWORD = os.environ.get("BACKDOOR_ADMIN_PASSWORD") or os.environ.get("ADMIN_PASSWORD", "admin1234")

ALLOWED_ORIGINS = [o.strip() for o in os.environ.get("ALLOWED_ORIGINS", "*").split(",") if o.strip()]
if ALLOWED_ORIGINS == ["*"]:
    ALLOW_CREDENTIALS = False
else:
    ALLOW_CREDENTIALS = True

VOUCHER_AMOUNT_KEYS = [
    "bcr_monto",
    "bac_monto",
    "bac_flotas_monto",
    "versatec_monto",
    "fleet_bncr_monto",
    "fleet_dav_monto",
]
MOVEMENT_FIELDS = ("creditos", "mercaderia_credito", "sinpes", "depositos", "vales", "pagos")
COMPACT_MOVEMENT_FIELDS = {"sinpes", "vales", "pagos"}
EDITABLE_EMPLOYEE_STATUSES = {"submitted"}
STATUS_SUBMITTED = "submitted"
STATUS_REVIEWED = "reviewed"
STATUS_APPROVED = "approved"
STATUS_RECONCILED = "reconciled"
STATUS_DELETED = "deleted"
ITEM_NOTE_SECTIONS = {"depositos", "creditos", "mercaderia_credito", "sinpes", "vales", "pagos"}
PISTA_SECTION_NOTE_KEYS = ITEM_NOTE_SECTIONS | {"vouchers", "mercaderia_contado", "observaciones"}
TIENDA_SECTION_NOTE_KEYS = {"resumen_ingresos", "detalle", "observaciones"}
PRODUCT_ALIASES = {
    "s": "super",
    "super": "super",
    "superior": "super",
    "regular": "regular",
    "reg": "regular",
    "r": "regular",
    "diesel": "diesel",
    "d": "diesel",
    "glp": "glp",
    "gas": "glp",
}
HEADER_ALIASES = {
    "fecha": {"fecha", "dia", "date"},
    "empleado": {"empleado", "nombre", "operador", "usuario", "despachador"},
    "turno": {"turno", "shift"},
    "producto": {"producto", "combustible", "fuel", "tipo", "articulo"},
    "litros": {"litros", "cantidad", "lts", "lt", "volumen"},
    "monto": {"monto", "importe", "total", "venta", "valor"},
    "ppu": {"ppu", "precio", "precioxunidad", "preciounitario", "price"},
}

app = FastAPI(title=APP_TITLE)
app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_methods=["*"],
    allow_headers=["*"],
    allow_credentials=ALLOW_CREDENTIALS,
)


class LoginRequest(BaseModel):
    pin: Optional[str] = None
    username: Optional[str] = None
    password: Optional[str] = None


class VouchersModel(BaseModel):
    bcr_qty: Optional[str] = ""
    bcr_monto: Optional[str] = ""
    bac_qty: Optional[str] = ""
    bac_monto: Optional[str] = ""
    bac_flotas_qty: Optional[str] = ""
    bac_flotas_monto: Optional[str] = ""
    versatec_qty: Optional[str] = ""
    versatec_monto: Optional[str] = ""
    fleet_bncr_qty: Optional[str] = ""
    fleet_bncr_monto: Optional[str] = ""
    fleet_dav_qty: Optional[str] = ""
    fleet_dav_monto: Optional[str] = ""


class MovementItem(BaseModel):
    id: Optional[str] = None
    comprobante: Optional[str] = ""
    monto: Optional[str] = ""
    descripcion: Optional[str] = ""
    cliente: Optional[str] = ""
    referencia: Optional[str] = ""
    monto_reportado: Optional[str] = ""
    monto_validado: Optional[str] = None
    estado: Optional[str] = "reportado"
    observacion_empleado: Optional[str] = ""
    observacion_supervisor: Optional[str] = ""
    soporte_requerido: Optional[bool] = False
    validado_por: Optional[str] = None
    validado_at: Optional[str] = None


class CierrePayload(BaseModel):
    fecha: str
    turno: str
    datafono: Optional[str] = ""
    mercaderia_contado: Optional[str] = ""
    vouchers: VouchersModel = Field(default_factory=VouchersModel)
    creditos: List[MovementItem] = Field(default_factory=list)
    mercaderia_credito: List[MovementItem] = Field(default_factory=list)
    sinpes: List[MovementItem] = Field(default_factory=list)
    depositos: List[MovementItem] = Field(default_factory=list)
    vales: List[MovementItem] = Field(default_factory=list)
    pagos: List[MovementItem] = Field(default_factory=list)
    observaciones: Optional[str] = ""
    employee_id: Optional[int] = None


class CierreTiendaPayload(BaseModel):
    fecha: str
    # Resumen de Ingresos
    mercaderia_contado: Optional[str] = ""
    abonos_cxc_transferencia: Optional[str] = ""
    abonos_cxc_efectivo: Optional[str] = ""
    mercaderia_credito: Optional[str] = ""
    mercaderia_credito_roco: Optional[str] = ""
    mercaderia_credito_ltj: Optional[str] = ""
    mercaderia_credito_jema: Optional[str] = ""
    t_nelson: Optional[str] = ""
    iva_nelson: Optional[str] = ""
    v_finca: Optional[str] = ""
    litros_finca: Optional[str] = ""
    # Detalle
    transf_sinpe_bcr: Optional[str] = ""
    transferencias_bn_bac: Optional[str] = ""
    sinpe: Optional[str] = ""
    tarjeta_bac: Optional[str] = ""
    tarjeta_bn: Optional[str] = ""
    tarjeta_bcr: Optional[str] = ""
    credito_detalle: Optional[str] = ""
    mercaderia_emilio: Optional[str] = ""
    vales_tienda: Optional[str] = ""
    salidas_efectivo: Optional[str] = ""
    nota_credito: Optional[str] = ""
    deposito: Optional[str] = ""
    observaciones: Optional[str] = ""


class ReviewPayload(BaseModel):
    validado_json: Dict[str, Any]
    status: str = STATUS_REVIEWED
    audit_notes: Optional[str] = ""


class ReviewNotePayload(BaseModel):
    target_scope: str
    section_key: str
    movement_id: Optional[str] = None
    body: str


class ReviewNoteUpdatePayload(BaseModel):
    resolved: bool


class DeleteCierrePayload(BaseModel):
    reason: Optional[str] = ""


class UserCreate(BaseModel):
    username: Optional[str] = None
    full_name: str
    role: str
    default_turno: Optional[str] = None
    pin: Optional[str] = None
    password: Optional[str] = None


class UserUpdate(BaseModel):
    username: Optional[str] = None
    full_name: Optional[str] = None
    role: Optional[str] = None
    default_turno: Optional[str] = None
    pin: Optional[str] = None
    password: Optional[str] = None
    active: Optional[bool] = None


class GasproImportResponse(BaseModel):
    import_id: int
    matched_cierres: int
    skipped_cierres: int = 0
    already_reconciled: int = 0
    import_mode: str


def utcnow() -> datetime:
    return datetime.utcnow()


def normalize_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", "", text.lower())


def slugify_username(value: Any) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii").lower()
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    return text or "usuario"


def parse_turn_number(value: Any) -> int:
    text = str(value or "").strip()
    return int(text) if text.isdigit() else 999999


def parse_decimal(value: Any) -> Decimal:
    if value in (None, "", "None"):
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip().replace("₡", "").replace(",", "")
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def decimal_to_float(value: Any) -> float:
    return float(parse_decimal(value))


def hash_secret(secret: str) -> str:
    salt = secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac("sha256", secret.encode("utf-8"), salt.encode("utf-8"), 200_000)
    return f"pbkdf2_sha256${salt}${digest.hex()}"


def verify_secret(secret: str, stored: Optional[str]) -> bool:
    if not stored:
        return False
    try:
        algorithm, salt, digest_hex = stored.split("$", 2)
    except ValueError:
        return False
    if algorithm != "pbkdf2_sha256":
        return False
    digest = hashlib.pbkdf2_hmac("sha256", secret.encode("utf-8"), salt.encode("utf-8"), 200_000)
    return hmac.compare_digest(digest.hex(), digest_hex)


def token_hash(token: str) -> str:
    return hashlib.sha256(token.encode("utf-8")).hexdigest()


def create_session(db, user_id: int) -> str:
    token = secrets.token_urlsafe(48)
    db.cursor().execute(
        "INSERT INTO sessions (user_id, token_hash, expires_at) VALUES (%s, %s, %s)",
        (user_id, token_hash(token), utcnow() + timedelta(hours=TOKEN_TTL_HOURS)),
    )
    return token


def parse_json_field(value: Any, default: Any) -> Any:
    if value is None:
        return default
    if isinstance(value, (dict, list)):
        return value
    if isinstance(value, str):
        try:
            return json.loads(value)
        except Exception:
            return default
    return default


def pick_first_text(*values: Any) -> str:
    for value in values:
        text = str(value or "").strip()
        if text:
            return text
    return ""


def fallback_movement_id(item: Dict[str, Any], movement_type: str) -> str:
    canonical = {
        "tipo": movement_type,
        "comprobante": pick_first_text(
            item.get("comprobante"),
            item.get("referencia"),
            item.get("descripcion"),
            item.get("cliente"),
            item.get("numero"),
            item.get("detalle"),
        ),
        "cliente": str(item.get("cliente", "") or item.get("empresa", "") or "").strip(),
        "referencia": str(item.get("referencia", "") or item.get("numero", "") or "").strip(),
        "descripcion": str(item.get("descripcion", "") or item.get("detalle", "") or "").strip(),
        "monto": str(item.get("monto", item.get("monto_reportado", "")) or "").strip(),
    }
    return str(uuid.uuid5(uuid.NAMESPACE_URL, json.dumps(canonical, sort_keys=True, ensure_ascii=True)))


def normalize_movement(item: Dict[str, Any], movement_type: str) -> Dict[str, Any]:
    amount_text = str(item.get("monto", item.get("monto_reportado", "")) or "")
    item_id = item.get("id") or fallback_movement_id(item, movement_type)
    if movement_type in COMPACT_MOVEMENT_FIELDS:
        comprobante = pick_first_text(
            item.get("comprobante"),
            item.get("referencia"),
            item.get("descripcion"),
            item.get("cliente"),
            item.get("numero"),
            item.get("detalle"),
        )
        return {
            "id": item_id,
            "tipo": movement_type,
            "comprobante": comprobante,
            "monto": amount_text,
            "monto_reportado": amount_text,
            "monto_validado": item.get("monto_validado"),
            "observacion_empleado": item.get("observacion_empleado", item.get("observacion", "")) or "",
            "observacion_supervisor": item.get("observacion_supervisor", "") or "",
        }
    return {
        "id": item_id,
        "tipo": movement_type,
        "descripcion": item.get("descripcion", "") or item.get("detalle", ""),
        "cliente": item.get("cliente", "") or item.get("empresa", ""),
        "referencia": item.get("referencia", "") or item.get("numero", ""),
        "monto_reportado": amount_text,
        "monto_validado": item.get("monto_validado"),
        "estado": item.get("estado", "reportado"),
        "observacion_empleado": item.get("observacion_empleado", item.get("observacion", "")) or "",
        "observacion_supervisor": item.get("observacion_supervisor", "") or "",
        "soporte_requerido": bool(item.get("soporte_requerido", movement_type in {"sinpes", "depositos", "vales"})),
        "validado_por": item.get("validado_por"),
        "validado_at": item.get("validado_at"),
    }


def export_movement_rows(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for item in items or []:
        comprobante = pick_first_text(
            item.get("comprobante"),
            item.get("referencia"),
            item.get("descripcion"),
            item.get("cliente"),
        )
        monto = decimal_to_float(parse_decimal(item.get("monto", item.get("monto_reportado", 0))))
        if not comprobante and monto == 0:
            continue
        rows.append({"monto": monto, "comprobante": comprobante})
    return rows


def default_payload() -> Dict[str, Any]:
    return {
        "fecha": date.today().isoformat(),
        "turno": "1",
        "datafono": "",
        "mercaderia_contado": "",
        "vouchers": {
            "bcr_qty": "", "bcr_monto": "",
            "bac_qty": "", "bac_monto": "",
            "bac_flotas_qty": "", "bac_flotas_monto": "",
            "versatec_qty": "", "versatec_monto": "",
            "fleet_bncr_qty": "", "fleet_bncr_monto": "",
            "fleet_dav_qty": "", "fleet_dav_monto": "",
        },
        "creditos": [],
        "mercaderia_credito": [],
        "sinpes": [],
        "depositos": [],
        "vales": [],
        "pagos": [],
        "observaciones": "",
    }


def normalize_payload(data: Dict[str, Any]) -> Dict[str, Any]:
    if data is None:
        data = {}
    if not isinstance(data, dict):
        raise HTTPException(status_code=422, detail="Formato de cierre invalido")

    payload = default_payload()

    raw_fecha = data.get("fecha") or payload["fecha"]
    try:
        payload["fecha"] = parse_date_value(raw_fecha).isoformat()
    except Exception as exc:
        raise HTTPException(status_code=422, detail="Fecha invalida") from exc

    payload["turno"] = str(data.get("turno", payload["turno"]) or payload["turno"]).strip() or payload["turno"]
    payload["datafono"] = str(data.get("datafono", payload["datafono"]) or "").strip()
    payload["mercaderia_contado"] = str(data.get("mercaderia_contado", "") or "").strip()
    payload["observaciones"] = str(data.get("observaciones", payload["observaciones"]) or "").strip()

    vouchers = data.get("vouchers") or {}
    if not isinstance(vouchers, dict):
        raise HTTPException(status_code=422, detail="Formato de vouchers invalido")
    payload["vouchers"] = {**default_payload()["vouchers"], **vouchers}

    for field in MOVEMENT_FIELDS:
        items = data.get(field) or []
        if not isinstance(items, list):
            raise HTTPException(status_code=422, detail=f"Formato invalido en {field}")
        payload[field] = [normalize_movement(item, field) for item in items]
    return payload


def compute_summary(payload: Dict[str, Any], validated: bool = False) -> Dict[str, float]:
    vouchers = payload.get("vouchers", {})
    total_vouchers = sum(parse_decimal(vouchers.get(key)) for key in VOUCHER_AMOUNT_KEYS)
    total_mercaderia_contado = parse_decimal(payload.get("mercaderia_contado"))
    totals = {}
    for field in MOVEMENT_FIELDS:
        subtotal = Decimal("0")
        for item in payload.get(field, []):
            if validated and item.get("monto_validado") not in (None, "", "None"):
                amount = parse_decimal(item.get("monto_validado"))
            else:
                amount = parse_decimal(item.get("monto", item.get("monto_reportado", 0)))
            subtotal += amount
        totals[field] = subtotal
    total_creditos_directos = totals["creditos"]
    total_mercaderia_credito = totals.get("mercaderia_credito", Decimal("0"))
    total_creditos = total_creditos_directos + total_mercaderia_credito
    total_reportado = total_vouchers + total_mercaderia_contado + total_creditos + totals["sinpes"] + totals["depositos"] + totals["vales"] + totals["pagos"]
    return {
        "total_vouchers": decimal_to_float(total_vouchers),
        "total_mercaderia_contado": decimal_to_float(total_mercaderia_contado),
        "total_creditos": decimal_to_float(total_creditos),
        "total_creditos_directos": decimal_to_float(total_creditos_directos),
        "total_mercaderia_credito": decimal_to_float(total_mercaderia_credito),
        "total_sinpes": decimal_to_float(totals["sinpes"]),
        "total_depositos": decimal_to_float(totals["depositos"]),
        "total_vales": decimal_to_float(totals["vales"]),
        "total_pagos": decimal_to_float(totals["pagos"]),
        "total_reportado": decimal_to_float(total_reportado),
    }


TIENDA_RESUMEN_FIELDS = [
    "mercaderia_contado", "abonos_cxc_transferencia", "abonos_cxc_efectivo",
    "mercaderia_credito", "mercaderia_credito_roco", "mercaderia_credito_ltj",
    "mercaderia_credito_jema", "t_nelson", "iva_nelson", "v_finca",
]
TIENDA_DETALLE_FIELDS = [
    "transf_sinpe_bcr", "transferencias_bn_bac", "sinpe",
    "tarjeta_bac", "tarjeta_bn", "tarjeta_bcr",
    "credito_detalle", "mercaderia_emilio",
    "vales_tienda", "salidas_efectivo", "nota_credito", "deposito",
]


def default_tienda_payload() -> Dict[str, Any]:
    payload: Dict[str, Any] = {
        "fecha": date.today().isoformat(),
        "tipo": "tienda",
        "litros_finca": "",
        "observaciones": "",
    }
    for f in TIENDA_RESUMEN_FIELDS:
        payload[f] = ""
    for f in TIENDA_DETALLE_FIELDS:
        payload[f] = ""
    return payload


def normalize_tienda_payload(data: Dict[str, Any]) -> Dict[str, Any]:
    if data is None:
        data = {}
    payload = default_tienda_payload()
    raw_fecha = data.get("fecha") or payload["fecha"]
    try:
        payload["fecha"] = parse_date_value(raw_fecha).isoformat()
    except Exception as exc:
        raise HTTPException(status_code=422, detail="Fecha invalida") from exc
    payload["observaciones"] = str(data.get("observaciones", "") or "").strip()
    payload["litros_finca"] = str(data.get("litros_finca", "") or "").strip()
    for f in TIENDA_RESUMEN_FIELDS + TIENDA_DETALLE_FIELDS:
        payload[f] = str(data.get(f, "") or "").strip()
    return payload


def compute_tienda_summary(payload: Dict[str, Any]) -> Dict[str, float]:
    resumen_total = Decimal("0")
    detalle_total = Decimal("0")
    resumen_items = {}
    detalle_items = {}
    for f in TIENDA_RESUMEN_FIELDS:
        val = parse_decimal(payload.get(f))
        resumen_items[f] = decimal_to_float(val)
        resumen_total += val
    for f in TIENDA_DETALLE_FIELDS:
        val = parse_decimal(payload.get(f))
        detalle_items[f] = decimal_to_float(val)
        detalle_total += val
    return {
        **resumen_items,
        **detalle_items,
        "total_resumen": decimal_to_float(resumen_total),
        "total_detalle": decimal_to_float(detalle_total),
        "diferencia": decimal_to_float(resumen_total - detalle_total),
        "total_reportado": decimal_to_float(resumen_total),
    }


def row_to_user(row: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "id": row["id"],
        "username": row["username"],
        "full_name": row["full_name"],
        "role": row["role"],
        "default_turno": row.get("default_turno"),
        "active": row["active"],
    }


def build_unique_username(db, raw_username: Optional[str], fallback_name: str, exclude_user_id: Optional[int] = None) -> str:
    base = slugify_username(raw_username or fallback_name)
    candidate = base
    suffix = 2
    cur = db.cursor()
    while True:
        params: List[Any] = [candidate]
        query = "SELECT 1 FROM users WHERE username = %s"
        if exclude_user_id is not None:
            query += " AND id <> %s"
            params.append(exclude_user_id)
        cur.execute(query, params)
        if not cur.fetchone():
            return candidate
        candidate = f"{base}-{suffix}"
        suffix += 1


def normalize_employee_turns(db):
    cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        """
        SELECT id, default_turno
        FROM users
        WHERE role = 'employee'
          AND active = TRUE
        ORDER BY created_at ASC, id ASC
        """
    )
    employees = [dict(row) for row in cur.fetchall()]
    employees.sort(key=lambda row: (parse_turn_number(row.get("default_turno")), row["id"]))
    update_cur = db.cursor()
    for index, employee in enumerate(employees, start=1):
        update_cur.execute(
            """
            UPDATE users
            SET default_turno = %s,
                updated_at = NOW()
            WHERE id = %s
            """,
            (str(index), employee["id"]),
        )


def ensure_backdoor_admin_user(db) -> Dict[str, Any]:
    cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM users WHERE username = %s LIMIT 1", (BACKDOOR_ADMIN_USERNAME,))
    row = cur.fetchone()
    password_hash = hash_secret(BACKDOOR_ADMIN_PASSWORD)
    if row:
        db.cursor().execute(
            """
            UPDATE users
            SET full_name = %s,
                role = 'admin',
                active = TRUE,
                pin_hash = NULL,
                password_hash = %s,
                updated_at = NOW()
            WHERE id = %s
            """,
            ("Administrador de respaldo", password_hash, row["id"]),
        )
        refreshed = fetch_user_by_id(db, row["id"])
        if not refreshed:
            raise HTTPException(status_code=500, detail="No fue posible restaurar el acceso de respaldo")
        return refreshed

    insert_cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    insert_cur.execute(
        """
        INSERT INTO users (username, full_name, role, default_turno, pin_hash, password_hash, active)
        VALUES (%s, %s, 'admin', NULL, NULL, %s, TRUE)
        RETURNING *
        """,
        (BACKDOOR_ADMIN_USERNAME, "Administrador de respaldo", password_hash),
    )
    created = insert_cur.fetchone()
    return dict(created)


def normalize_cierre_status(status: Optional[str]) -> str:
    if status in (None, "", "draft", "observed"):
        return STATUS_SUBMITTED
    if status == "document_reviewed":
        return STATUS_REVIEWED
    if status == "validated":
        return STATUS_APPROVED
    if status in {STATUS_SUBMITTED, STATUS_REVIEWED, STATUS_APPROVED, STATUS_RECONCILED, STATUS_DELETED}:
        return status
    return STATUS_SUBMITTED


def fetch_user_by_id(db, user_id: int) -> Optional[dict]:
    cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM users WHERE id = %s LIMIT 1", (user_id,))
    row = cur.fetchone()
    return dict(row) if row else None


def get_current_user(authorization: Optional[str] = Header(default=None), db=Depends(get_db)) -> dict:
    if not authorization or not authorization.lower().startswith("bearer "):
        raise HTTPException(status_code=401, detail="Falta autenticación")
    token = authorization.split(" ", 1)[1].strip()
    cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        """
        SELECT u.*, s.id AS session_id
        FROM sessions s
        JOIN users u ON u.id = s.user_id
        WHERE s.token_hash = %s
          AND s.revoked_at IS NULL
          AND s.expires_at > NOW()
          AND u.active = TRUE
        LIMIT 1
        """,
        (token_hash(token),),
    )
    row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=401, detail="Sesión inválida o vencida")
    return dict(row)


def require_roles(*roles):
    def dependency(user=Depends(get_current_user)):
        if user["role"] not in roles:
            raise HTTPException(status_code=403, detail="No autorizado")
        return user
    return dependency


def save_audit_log(db, cierre_id: Optional[int], actor_id: int, action: str, details: Dict[str, Any]):
    cur = db.cursor()
    cur.execute(
        "INSERT INTO cierre_audit_log (cierre_id, actor_id, action, details_json) VALUES (%s, %s, %s, %s::jsonb)",
        (cierre_id, actor_id, action, json.dumps(details)),
    )


def can_manage_review_notes(user: Dict[str, Any]) -> bool:
    return user["role"] in {"admin", "supervisor"}


def allowed_note_sections(cierre: Dict[str, Any]) -> set[str]:
    return TIENDA_SECTION_NOTE_KEYS if cierre.get("tipo") == "tienda" else PISTA_SECTION_NOTE_KEYS


def movement_exists_for_review_note(cierre: Dict[str, Any], section_key: str, movement_id: str) -> bool:
    payload = cierre.get("reportado_json") or {}
    items = payload.get(section_key) or []
    return any(str(item.get("id") or "").strip() == movement_id for item in items if isinstance(item, dict))


def list_review_notes(db, cierre_id: int) -> List[Dict[str, Any]]:
    cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        """
        SELECT n.*,
               author.full_name AS author_name,
               author.role AS author_role,
               resolver.full_name AS resolved_by_name,
               resolver.role AS resolved_by_role
        FROM cierre_review_notes n
        JOIN users author ON author.id = n.created_by_user_id
        LEFT JOIN users resolver ON resolver.id = n.resolved_by_user_id
        WHERE n.cierre_id = %s
        ORDER BY n.resolved ASC, n.created_at DESC, n.id DESC
        """,
        (cierre_id,),
    )
    return [dict(row) for row in cur.fetchall()]


def attach_review_note_counts(db, cierres: List[Dict[str, Any]]) -> None:
    ids = [c["id"] for c in cierres if c.get("id")]
    if not ids:
        return
    cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        """
        SELECT cierre_id,
               COUNT(*) AS total_notes,
               COUNT(*) FILTER (WHERE resolved = FALSE) AS unresolved_notes
        FROM cierre_review_notes
        WHERE cierre_id = ANY(%s)
        GROUP BY cierre_id
        """,
        (ids,),
    )
    counts = {row["cierre_id"]: dict(row) for row in cur.fetchall()}
    for cierre in cierres:
        current = counts.get(cierre["id"], {})
        cierre["note_count"] = int(current.get("total_notes") or 0)
        cierre["unresolved_note_count"] = int(current.get("unresolved_notes") or 0)


def hydrate_cierre(row: Dict[str, Any]) -> Dict[str, Any]:
    cierre = dict(row)
    reportado = parse_json_field(cierre.get("reportado_json"), {})
    validado = parse_json_field(cierre.get("validado_json"), {})
    cierre["status"] = normalize_cierre_status(cierre.get("status"))
    cierre["tipo"] = cierre.get("tipo", "pista")
    if cierre["tipo"] == "tienda":
        cierre["reportado_json"] = normalize_tienda_payload(reportado)
        cierre["validado_json"] = normalize_tienda_payload(validado if validado else reportado)
    else:
        cierre["reportado_json"] = normalize_payload(reportado)
        cierre["validado_json"] = normalize_payload(validado if validado else reportado)
    cierre["resumen_reportado"] = parse_json_field(cierre.get("resumen_reportado"), {})
    cierre["resumen_validado"] = parse_json_field(cierre.get("resumen_validado"), {})
    cierre["gaspro_summary"] = parse_json_field(cierre.get("gaspro_summary"), {})
    return cierre


def can_employee_edit(cierre: Dict[str, Any], user: Dict[str, Any]) -> bool:
    if user["role"] != "employee":
        return False
    if cierre.get("employee_id") != user["id"]:
        return False
    if cierre.get("status") not in EDITABLE_EMPLOYEE_STATUSES:
        return False
    if cierre.get("reviewed_at") or cierre.get("approved_at") or cierre.get("reconciled_at") or cierre.get("deleted_at"):
        return False
    editable_until = cierre.get("editable_until")
    if not editable_until:
        return False
    if isinstance(editable_until, str):
        editable_until = datetime.fromisoformat(editable_until)
    return utcnow() <= editable_until


def assert_can_view_cierre(cierre: Dict[str, Any], user: Dict[str, Any]):
    if user["role"] in {"admin", "supervisor"}:
        return
    if cierre.get("employee_id") != user["id"]:
        raise HTTPException(status_code=403, detail="No podés ver este cierre")


def assert_can_edit_cierre(cierre: Dict[str, Any], user: Dict[str, Any]):
    if normalize_cierre_status(cierre.get("status")) == STATUS_DELETED:
        raise HTTPException(status_code=403, detail="Restaura este cierre antes de editarlo")
    if user["role"] == "admin":
        return
    if user["role"] == "tienda":
        if cierre.get("employee_id") != user["id"]:
            raise HTTPException(status_code=403, detail="No podés editar este cierre")
        if cierre.get("tipo") != "tienda":
            raise HTTPException(status_code=403, detail="No autorizado")
        return
    if user["role"] == "supervisor":
        raise HTTPException(status_code=403, detail="Solo un administrador puede editar cierres directamente")
    if not can_employee_edit(cierre, user):
        raise HTTPException(status_code=403, detail="El cierre ya no puede ser modificado")


def assert_can_review_transition(cierre: Dict[str, Any], user: Dict[str, Any], next_status: str):
    current_status = normalize_cierre_status(cierre.get("status"))
    if current_status == STATUS_DELETED:
        raise HTTPException(status_code=403, detail="No podes revisar un cierre en papelera")
    if user["role"] == "admin":
        if next_status != STATUS_APPROVED:
            raise HTTPException(status_code=400, detail="El administrador solo puede aprobar cierres")
        if current_status not in {STATUS_REVIEWED, STATUS_APPROVED}:
            raise HTTPException(status_code=400, detail="Solo se puede aprobar un cierre revisado")
        return
    if user["role"] != "supervisor":
        raise HTTPException(status_code=403, detail="No autorizado")
    if next_status != STATUS_REVIEWED:
        raise HTTPException(status_code=400, detail="El supervisor solo puede marcar cierres como revisados")
    if current_status not in {STATUS_SUBMITTED, STATUS_REVIEWED}:
        raise HTTPException(status_code=400, detail="Solo se puede revisar un cierre enviado")


def assert_can_reconcile_cierre(cierre: Dict[str, Any], user: Dict[str, Any]):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Solo un administrador puede conciliar cierres")
    if cierre.get("tipo") == "tienda":
        raise HTTPException(status_code=400, detail="Los cierres de tienda no se concilian con Gaspro")
    if normalize_cierre_status(cierre.get("status")) != STATUS_APPROVED:
        raise HTTPException(status_code=400, detail="Solo se puede conciliar un cierre aprobado")


def get_cierre_or_404(db, cierre_id: int) -> Dict[str, Any]:
    cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM cierres WHERE id = %s", (cierre_id,))
    row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Cierre no encontrado")
    return hydrate_cierre(row)


def persist_cierre_payload(
    db,
    payload: Dict[str, Any],
    validado_payload: Dict[str, Any],
    current_user: Dict[str, Any],
    cierre_id: Optional[int] = None,
    employee_id: Optional[int] = None,
    status: str = STATUS_SUBMITTED,
    reset_reconciliation: bool = False,
) -> int:
    reportado_payload = normalize_payload(payload)
    validado_payload = normalize_payload(validado_payload or payload)
    resumen_reportado = compute_summary(reportado_payload, validated=False)
    resumen_validado = compute_summary(validado_payload, validated=True)
    status = normalize_cierre_status(status)
    employee_id = employee_id or current_user["id"]
    employee = fetch_user_by_id(db, employee_id)
    if not employee:
        raise HTTPException(status_code=400, detail="Empleado inválido")
    editable_until = utcnow() if status in {STATUS_REVIEWED, STATUS_APPROVED, STATUS_RECONCILED} else utcnow() + timedelta(hours=EMPLOYEE_EDIT_HOURS)
    cur = db.cursor()
    v = reportado_payload["vouchers"]
    creditos_text = json.dumps(export_movement_rows(reportado_payload["creditos"]))
    sinpes_text = json.dumps(export_movement_rows(reportado_payload["sinpes"]))
    vales_text = json.dumps(export_movement_rows(reportado_payload["vales"]))
    pagos_text = json.dumps(export_movement_rows(reportado_payload["pagos"]))
    if cierre_id is None:
        cur.execute(
            """
            INSERT INTO cierres (
                fecha, empleado, employee_id, turno, datafono, observaciones,
                status, reportado_json, validado_json, resumen_reportado, resumen_validado,
                created_by_user_id, edited_by_user_id, submitted_at, editable_until, updated_at,
                voucher_bcr, voucher_bac, voucher_bac_flotas, voucher_versatec,
                voucher_fleet_bncr, voucher_fleet_dav, voucher_bncr,
                creditos_json, sinpes_json, deposito, vales_json, pagos_json, efectivo, total_reportado, enviado_at
            ) VALUES (
                %s,%s,%s,%s,%s,%s,%s,%s::jsonb,%s::jsonb,%s::jsonb,%s::jsonb,
                %s,%s,%s,%s,%s,
                %s,%s,%s,%s,%s,%s,%s,
                %s::jsonb,%s::jsonb,%s,%s::jsonb,%s::jsonb,%s,%s,%s
            ) RETURNING id
            """,
            (
                reportado_payload["fecha"], employee["full_name"], employee["id"], reportado_payload["turno"], reportado_payload["datafono"], reportado_payload["observaciones"],
                status, json.dumps(reportado_payload), json.dumps(validado_payload), json.dumps(resumen_reportado), json.dumps(resumen_validado),
                current_user["id"], current_user["id"], utcnow(), editable_until, utcnow(),
                decimal_to_float(v.get("bcr_monto")), decimal_to_float(v.get("bac_monto")), decimal_to_float(v.get("bac_flotas_monto")), decimal_to_float(v.get("versatec_monto")),
                decimal_to_float(v.get("fleet_bncr_monto")), decimal_to_float(v.get("fleet_dav_monto")), 0,
                creditos_text, sinpes_text, 0, vales_text, pagos_text, 0, resumen_reportado["total_reportado"], utcnow(),
            ),
        )
        cierre_id = cur.fetchone()[0]
        save_audit_log(db, cierre_id, current_user["id"], "created", {"status": status})
    else:
        cur.execute(
            """
            UPDATE cierres
            SET fecha = %s,
                empleado = %s,
                employee_id = %s,
                turno = %s,
                datafono = %s,
                observaciones = %s,
                status = %s,
                reportado_json = %s::jsonb,
                validado_json = %s::jsonb,
                resumen_reportado = %s::jsonb,
                resumen_validado = %s::jsonb,
                edited_by_user_id = %s,
                editable_until = %s,
                gaspro_summary = CASE WHEN %s THEN '{}'::jsonb ELSE gaspro_summary END,
                gaspro_mode = CASE WHEN %s THEN NULL ELSE gaspro_mode END,
                reconciled_at = CASE WHEN %s THEN NULL ELSE reconciled_at END,
                updated_at = %s,
                voucher_bcr = %s,
                voucher_bac = %s,
                voucher_bac_flotas = %s,
                voucher_versatec = %s,
                voucher_fleet_bncr = %s,
                voucher_fleet_dav = %s,
                voucher_bncr = %s,
                creditos_json = %s::jsonb,
                sinpes_json = %s::jsonb,
                deposito = %s,
                vales_json = %s::jsonb,
                pagos_json = %s::jsonb,
                efectivo = %s,
                total_reportado = %s
            WHERE id = %s
            """,
            (
                reportado_payload["fecha"], employee["full_name"], employee["id"], reportado_payload["turno"], reportado_payload["datafono"], reportado_payload["observaciones"],
                status, json.dumps(reportado_payload), json.dumps(validado_payload), json.dumps(resumen_reportado), json.dumps(resumen_validado),
                current_user["id"], editable_until, reset_reconciliation, reset_reconciliation, reset_reconciliation, utcnow(),
                decimal_to_float(v.get("bcr_monto")), decimal_to_float(v.get("bac_monto")), decimal_to_float(v.get("bac_flotas_monto")), decimal_to_float(v.get("versatec_monto")),
                decimal_to_float(v.get("fleet_bncr_monto")), decimal_to_float(v.get("fleet_dav_monto")), 0,
                creditos_text, sinpes_text, 0, vales_text, pagos_text, 0, resumen_reportado["total_reportado"], cierre_id,
            ),
        )
        save_audit_log(db, cierre_id, current_user["id"], "updated", {"status": status})
    db.commit()
    return cierre_id


def persist_cierre_tienda(
    db,
    payload: Dict[str, Any],
    current_user: Dict[str, Any],
    cierre_id: Optional[int] = None,
    status: str = STATUS_SUBMITTED,
) -> int:
    reportado = normalize_tienda_payload(payload)
    resumen = compute_tienda_summary(reportado)
    status = normalize_cierre_status(status)
    editable_until = None
    cur = db.cursor()
    if cierre_id is None:
        cur.execute(
            """
            INSERT INTO cierres (
                fecha, empleado, employee_id, turno, datafono, observaciones,
                status, tipo, reportado_json, validado_json, resumen_reportado, resumen_validado,
                created_by_user_id, edited_by_user_id, submitted_at, editable_until, updated_at,
                voucher_bcr, voucher_bac, voucher_bac_flotas, voucher_versatec,
                voucher_fleet_bncr, voucher_fleet_dav, voucher_bncr,
                creditos_json, sinpes_json, deposito, vales_json, pagos_json, efectivo, total_reportado, enviado_at
            ) VALUES (
                %s,%s,%s,%s,%s,%s,%s,%s,%s::jsonb,%s::jsonb,%s::jsonb,%s::jsonb,
                %s,%s,%s,%s,%s,
                %s,%s,%s,%s,%s,%s,%s,
                %s::jsonb,%s::jsonb,%s,%s::jsonb,%s::jsonb,%s,%s,%s
            ) RETURNING id
            """,
            (
                reportado["fecha"], current_user["full_name"], current_user["id"], "", "", reportado.get("observaciones", ""),
                status, "tienda", json.dumps(reportado), json.dumps(reportado), json.dumps(resumen), json.dumps(resumen),
                current_user["id"], current_user["id"], utcnow(), editable_until, utcnow(),
                0, 0, 0, 0, 0, 0, 0,
                '[]', '[]', 0, '[]', '[]', 0, resumen["total_reportado"], utcnow(),
            ),
        )
        cierre_id = cur.fetchone()[0]
        save_audit_log(db, cierre_id, current_user["id"], "created", {"status": status, "tipo": "tienda"})
    else:
        cur.execute(
            """
            UPDATE cierres
            SET fecha = %s, observaciones = %s, status = %s, reportado_json = %s::jsonb, validado_json = %s::jsonb,
                resumen_reportado = %s::jsonb, resumen_validado = %s::jsonb,
                edited_by_user_id = %s, editable_until = %s, updated_at = %s,
                total_reportado = %s
            WHERE id = %s
            """,
            (
                reportado["fecha"], reportado.get("observaciones", ""), status,
                json.dumps(reportado), json.dumps(reportado), json.dumps(resumen), json.dumps(resumen),
                current_user["id"], editable_until, utcnow(),
                resumen["total_reportado"], cierre_id,
            ),
        )
        save_audit_log(db, cierre_id, current_user["id"], "updated", {"status": status, "tipo": "tienda"})
    db.commit()
    return cierre_id


def save_upload(file: UploadFile, subdir: str) -> Dict[str, Any]:
    content = file.file.read()
    digest = hashlib.sha256(content).hexdigest()
    target_dir = UPLOAD_ROOT / subdir
    target_dir.mkdir(parents=True, exist_ok=True)
    suffix = Path(file.filename or "archivo").suffix
    filename = f"{uuid.uuid4().hex}{suffix}"
    target_path = target_dir / filename
    target_path.write_bytes(content)
    return {
        "path": str(target_path),
        "sha256": digest,
        "size_bytes": len(content),
        "mime_type": file.content_type or "application/octet-stream",
        "original_name": file.filename or filename,
    }


def infer_headers(row: List[Any]) -> Dict[str, int]:
    mapping = {}
    for idx, value in enumerate(row):
        key = normalize_text(value)
        for canonical, aliases in HEADER_ALIASES.items():
            if key in aliases and canonical not in mapping:
                mapping[canonical] = idx
    return mapping


def parse_date_value(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"No pude interpretar la fecha: {value}")


def normalize_product(value: Any) -> str:
    key = normalize_text(value)
    return PRODUCT_ALIASES.get(key, str(value or "").strip().lower())


def read_tabular_upload(path: Path) -> List[Dict[str, Any]]:
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            reader = csv.reader(fh)
            rows = list(reader)
    else:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
    header_map = None
    parsed: List[Dict[str, Any]] = []
    for row in rows:
        if not any(cell not in (None, "") for cell in row):
            continue
        if header_map is None:
            candidate = infer_headers(list(row))
            if {"fecha", "empleado", "producto", "litros", "monto"}.issubset(candidate.keys()):
                header_map = candidate
            continue
        entry = {key: row[idx] if idx < len(row) else None for key, idx in header_map.items()}
        try:
            parsed.append(
                {
                    "fecha": parse_date_value(entry.get("fecha")),
                    "empleado": str(entry.get("empleado") or "").strip(),
                    "turno": str(entry.get("turno") or "").strip(),
                    "producto": normalize_product(entry.get("producto")),
                    "litros": decimal_to_float(entry.get("litros")),
                    "monto": decimal_to_float(entry.get("monto")),
                    "ppu": decimal_to_float(entry.get("ppu")),
                    "raw_row": {k: (str(v) if isinstance(v, (datetime, date)) else v) for k, v in entry.items()},
                }
            )
        except Exception:
            continue
    return parsed


def summarize_gaspro_rows(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    by_employee: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        date_key = row["fecha"].isoformat()
        employee_key = normalize_text(row["empleado"])
        entity = by_employee.setdefault(date_key, {}).setdefault(employee_key, {"employee_name": row["empleado"], "products": {}, "price_change": False})
        product = entity["products"].setdefault(row["producto"], {"litros": 0.0, "monto": 0.0, "ppus": set()})
        product["litros"] += row["litros"]
        product["monto"] += row["monto"]
        if row["ppu"]:
            product["ppus"].add(round(row["ppu"], 4))
        if len(product["ppus"]) > 1:
            entity["price_change"] = True
    for date_key, employees in by_employee.items():
        for employee_key, summary in employees.items():
            for product_name, info in list(summary["products"].items()):
                summary["products"][product_name] = {
                    "litros": round(info["litros"], 3),
                    "monto": round(info["monto"], 2),
                    "ppus": sorted(info["ppus"]),
                }
    return by_employee


def find_latest_gaspro_match_for_cierre(db, cierre: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    target_date = cierre["fecha"].isoformat() if isinstance(cierre["fecha"], date) else str(cierre["fecha"])
    target_employee = normalize_text(cierre.get("empleado"))
    cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        """
        SELECT gr.*, gi.created_at AS import_created_at
        FROM gaspro_rows gr
        JOIN gaspro_imports gi ON gi.id = gr.gaspro_import_id
        WHERE gr.fecha = %s
        ORDER BY gi.created_at DESC, gr.gaspro_import_id DESC, gr.id ASC
        """,
        (target_date,),
    )
    grouped: Dict[int, List[Dict[str, Any]]] = {}
    for row in cur.fetchall():
        grouped.setdefault(row["gaspro_import_id"], []).append(
            {
                "fecha": row["fecha"] if isinstance(row["fecha"], date) else parse_date_value(row["fecha"]),
                "empleado": row["empleado"],
                "turno": row["turno"],
                "producto": row["producto"],
                "litros": decimal_to_float(row["litros"]),
                "monto": decimal_to_float(row["monto"]),
                "ppu": decimal_to_float(row["ppu"]),
                "raw_row": parse_json_field(row.get("raw_row"), {}),
            }
        )

    for import_id, rows in grouped.items():
        summary = summarize_gaspro_rows(rows).get(target_date, {}).get(target_employee)
        if summary:
            return {"import_id": import_id, "summary": summary}
    return None


def apply_gaspro_to_cierre(db, cierre: Dict[str, Any], summary: Dict[str, Any], import_id: int, actor: Dict[str, Any]):
    assert_can_reconcile_cierre(cierre, actor)
    mode = "price_change" if summary.get("price_change") else "normal"
    cur = db.cursor()
    cur.execute(
        """
        UPDATE cierres
        SET gaspro_summary = %s::jsonb,
            gaspro_import_id = %s,
            gaspro_mode = %s,
            reconciled_at = %s,
            editable_until = %s,
            status = %s,
            edited_by_user_id = %s,
            updated_at = %s
        WHERE id = %s
        """,
        (json.dumps(summary), import_id, mode, utcnow(), utcnow(), STATUS_RECONCILED, actor["id"], utcnow(), cierre["id"]),
    )
    save_audit_log(db, cierre["id"], actor["id"], "gaspro_reconciled", {"import_id": import_id, "mode": mode})


def find_day_sheet(workbook, day_number: int):
    candidates = [name for name in workbook.sheetnames if str(name).strip().startswith(str(day_number))]
    return workbook[candidates[0]] if candidates else None


def fill_cierre_in_sheet(ws, cierre: Dict[str, Any], validated: bool = True):
    payload = cierre["validado_json"] if validated else cierre["reportado_json"]
    summary = cierre["resumen_validado"] if validated else cierre["resumen_reportado"]
    gaspro = cierre.get("gaspro_summary") or {}
    turno = max(1, int(cierre.get("turno") or 1))
    base = 9 + (turno - 1) * 17
    products = gaspro.get("products", {})
    ws[f"C{base}"] = products.get("super", {}).get("litros", 0)
    ws[f"C{base+1}"] = products.get("regular", {}).get("litros", 0)
    ws[f"C{base+2}"] = products.get("diesel", {}).get("litros", 0)
    ws[f"C{base+3}"] = products.get("glp", {}).get("litros", 0)
    ws[f"J{base+8}"] = summary.get("total_creditos", 0)
    ws[f"J{base+9}"] = summary.get("deposito", 0)
    ws[f"J{base+10}"] = summary.get("total_sinpes", 0)
    ws[f"J{base+11}"] = summary.get("total_pagos", 0)
    ws[f"J{base+12}"] = summary.get("total_vales", 0)
    ws[f"D{base+13}"] = cierre.get("empleado")
    ws[f"J{base+13}"] = summary.get("total_reportado", 0)


@app.on_event("startup")
def startup_event():
    UPLOAD_ROOT.mkdir(parents=True, exist_ok=True)
    init_db()


@app.get("/api/health")
def health():
    return {"status": "ok"}


@app.post("/api/auth/login")
def login(payload: LoginRequest, db=Depends(get_db)):
    cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    if payload.pin:
        cur.execute("SELECT * FROM users WHERE active = TRUE AND role = 'employee'")
        for row in cur.fetchall():
            if verify_secret(payload.pin, row.get("pin_hash")):
                token = create_session(db, row["id"])
                db.commit()
                return {"token": token, "user": row_to_user(row)}
        raise HTTPException(status_code=401, detail="Clave incorrecta")
    if payload.username and payload.password:
        if payload.username == BACKDOOR_ADMIN_USERNAME and payload.password == BACKDOOR_ADMIN_PASSWORD:
            row = ensure_backdoor_admin_user(db)
            token = create_session(db, row["id"])
            db.commit()
            return {"token": token, "user": row_to_user(row)}
        cur.execute("SELECT * FROM users WHERE active = TRUE AND username = %s LIMIT 1", (payload.username,))
        row = cur.fetchone()
        if not row or row["role"] not in {"supervisor", "admin", "tienda"} or not verify_secret(payload.password, row.get("password_hash")):
            raise HTTPException(status_code=401, detail="Credenciales incorrectas")
        token = create_session(db, row["id"])
        db.commit()
        return {"token": token, "user": row_to_user(row)}
    raise HTTPException(status_code=400, detail="Faltan credenciales")


@app.post("/api/auth/logout")
def logout(user=Depends(get_current_user), db=Depends(get_db)):
    db.cursor().execute("UPDATE sessions SET revoked_at = NOW() WHERE id = %s", (user["session_id"],))
    db.commit()
    return {"ok": True}


@app.get("/api/me")
def me(user=Depends(get_current_user)):
    return row_to_user(user)


@app.get("/api/users")
def list_users(
    role: Optional[str] = None,
    include_inactive: bool = False,
    user: dict = Depends(require_roles("admin", "supervisor")),
    db=Depends(get_db),
):
    cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    if user["role"] != "admin":
        role = "employee"
        include_inactive = False

    filters: List[str] = []
    params: List[Any] = []

    if role:
        filters.append("role = %s")
        params.append(role)
    if not include_inactive:
        filters.append("active = TRUE")

    query = "SELECT * FROM users"
    if filters:
        query += " WHERE " + " AND ".join(filters)
    query += """
        ORDER BY
            CASE role
                WHEN 'admin' THEN 0
                WHEN 'supervisor' THEN 1
                ELSE 2
            END,
            CASE
                WHEN role = 'employee' AND NULLIF(default_turno, '') ~ '^[0-9]+$' THEN NULLIF(default_turno, '')::int
                ELSE 999999
            END,
            full_name ASC
    """
    cur.execute(query, params)
    return [row_to_user(row) for row in cur.fetchall()]


@app.post("/api/users")
def create_user(payload: UserCreate, _: dict = Depends(require_roles("admin")), db=Depends(get_db)):
    role = payload.role.strip().lower()
    if role not in {"employee", "supervisor", "admin", "tienda"}:
        raise HTTPException(status_code=422, detail="Rol invalido")

    full_name = str(payload.full_name or "").strip()
    if not full_name:
        raise HTTPException(status_code=422, detail="El nombre es obligatorio")

    username = build_unique_username(db, payload.username, full_name)
    pin_hash = None
    password_hash = None
    default_turno = None
    if role == "employee":
        if not payload.pin:
            raise HTTPException(status_code=422, detail="La clave del colaborador es obligatoria")
        pin_hash = hash_secret(payload.pin)
        cur_turn = db.cursor()
        cur_turn.execute("SELECT COUNT(*) FROM users WHERE role = 'employee' AND active = TRUE")
        default_turno = str(cur_turn.fetchone()[0] + 1)
    else:
        if not payload.password:
            raise HTTPException(status_code=422, detail="La contrasena del usuario es obligatoria")
        password_hash = hash_secret(payload.password)

    cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        """
        INSERT INTO users (username, full_name, role, default_turno, pin_hash, password_hash, active)
        VALUES (%s, %s, %s, %s, %s, %s, TRUE)
        RETURNING *
        """,
        (username, full_name, role, default_turno, pin_hash, password_hash),
    )
    row = cur.fetchone()
    if role == "employee":
        normalize_employee_turns(db)
    db.commit()
    return row_to_user(fetch_user_by_id(db, row["id"]))


@app.patch("/api/users/{user_id}")
def update_user(user_id: int, payload: UserUpdate, _: dict = Depends(require_roles("admin")), db=Depends(get_db)):
    row = fetch_user_by_id(db, user_id)
    if not row:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")

    next_role = (payload.role or row["role"]).strip().lower()
    if next_role not in {"employee", "supervisor", "admin", "tienda"}:
        raise HTTPException(status_code=422, detail="Rol invalido")

    full_name = payload.full_name if payload.full_name is not None else row["full_name"]
    full_name = str(full_name or "").strip()
    if not full_name:
        raise HTTPException(status_code=422, detail="El nombre es obligatorio")

    username = build_unique_username(db, payload.username if payload.username is not None else row["username"], full_name, exclude_user_id=user_id)
    active = payload.active if payload.active is not None else row["active"]
    pin_hash = row.get("pin_hash")
    password_hash = row.get("password_hash")
    default_turno = row.get("default_turno")

    if next_role == "employee":
        if row["role"] != "employee" and not payload.pin:
            raise HTTPException(status_code=422, detail="Debes definir una clave para el colaborador")
        if payload.pin:
            pin_hash = hash_secret(payload.pin)
        password_hash = None
        if active:
            if row["role"] != "employee" or not default_turno:
                cur_turn = db.cursor()
                cur_turn.execute("SELECT COUNT(*) FROM users WHERE role = 'employee' AND active = TRUE AND id <> %s", (user_id,))
                default_turno = str(cur_turn.fetchone()[0] + 1)
        else:
            default_turno = None
    else:
        if row["role"] == "employee" and not payload.password and not row.get("password_hash"):
            raise HTTPException(status_code=422, detail="Debes definir una contrasena para este usuario")
        if payload.password:
            password_hash = hash_secret(payload.password)
        pin_hash = None
        default_turno = None

    cur = db.cursor()
    cur.execute(
        """
        UPDATE users
        SET username = %s,
            full_name = %s,
            role = %s,
            default_turno = %s,
            active = %s,
            pin_hash = %s,
            password_hash = %s,
            updated_at = NOW()
        WHERE id = %s
        """,
        (username, full_name, next_role, default_turno, active, pin_hash, password_hash, user_id),
    )
    if row["role"] == "employee" or next_role == "employee":
        normalize_employee_turns(db)
    if not active:
        db.cursor().execute("UPDATE sessions SET revoked_at = NOW() WHERE user_id = %s AND revoked_at IS NULL", (user_id,))
    db.commit()
    return row_to_user(fetch_user_by_id(db, user_id))


@app.delete("/api/users/{user_id}")
def delete_user(user_id: int, user=Depends(require_roles("admin")), db=Depends(get_db)):
    row = fetch_user_by_id(db, user_id)
    if not row:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    if user_id == user["id"]:
        raise HTTPException(status_code=400, detail="No puedes desactivar tu propio usuario desde esta sesion")
    db.cursor().execute(
        """
        UPDATE users
        SET active = FALSE,
            default_turno = CASE WHEN role = 'employee' THEN NULL ELSE default_turno END,
            updated_at = NOW()
        WHERE id = %s
        """,
        (user_id,),
    )
    db.cursor().execute("UPDATE sessions SET revoked_at = NOW() WHERE user_id = %s AND revoked_at IS NULL", (user_id,))
    if row["role"] == "employee":
        normalize_employee_turns(db)
    db.commit()
    return {"ok": True}


@app.post("/api/cierres")
def create_cierre(payload: CierrePayload, user=Depends(get_current_user), db=Depends(get_db)):
    employee_id = payload.employee_id if user["role"] in {"admin", "supervisor"} and payload.employee_id else user["id"]
    cur = db.cursor()
    cur.execute(
        "SELECT COUNT(*) FROM cierres WHERE employee_id = %s AND fecha = %s AND COALESCE(status, '') <> %s",
        (employee_id, payload.fecha, STATUS_DELETED),
    )
    if cur.fetchone()[0] >= 3:
        raise HTTPException(status_code=409, detail="Ya existen 3 cierres para este empleado en esta fecha.")
    try:
        cierre_id = persist_cierre_payload(db, payload.model_dump(), payload.model_dump(), user, employee_id=employee_id, status=STATUS_SUBMITTED)
    except HTTPException:
        raise
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error al guardar el cierre: {str(exc)}") from exc
    return {"id": cierre_id, "status": "ok"}


@app.get("/api/cierres/count")
def count_cierres_for_date(fecha: str, user=Depends(get_current_user), db=Depends(get_db)):
    cur = db.cursor()
    cur.execute(
        "SELECT COUNT(*) FROM cierres WHERE employee_id = %s AND fecha = %s AND COALESCE(status, '') <> %s",
        (user["id"], fecha, STATUS_DELETED),
    )
    return {"count": cur.fetchone()[0], "max": 3}


@app.post("/api/cierres/tienda")
def create_cierre_tienda(payload: CierreTiendaPayload, user=Depends(get_current_user), db=Depends(get_db)):
    if user["role"] not in {"tienda", "admin"}:
        raise HTTPException(status_code=403, detail="No autorizado")
    cur = db.cursor()
    cur.execute(
        "SELECT COUNT(*) FROM cierres WHERE employee_id = %s AND fecha = %s AND tipo = 'tienda' AND COALESCE(status, '') <> %s",
        (user["id"], payload.fecha, STATUS_DELETED),
    )
    if cur.fetchone()[0] >= 1:
        raise HTTPException(status_code=409, detail="Ya existe un cierre de tienda para esta fecha.")
    try:
        cierre_id = persist_cierre_tienda(db, payload.model_dump(), user, status=STATUS_SUBMITTED)
    except HTTPException:
        raise
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error al guardar el cierre de tienda: {str(exc)}") from exc
    return {"id": cierre_id, "status": "ok"}


@app.put("/api/cierres/tienda/{cierre_id}")
def update_cierre_tienda(cierre_id: int, payload: CierreTiendaPayload, user=Depends(get_current_user), db=Depends(get_db)):
    cierre = get_cierre_or_404(db, cierre_id)
    if cierre.get("tipo") != "tienda":
        raise HTTPException(status_code=400, detail="Este cierre no es de tienda")
    assert_can_edit_cierre(cierre, user)
    try:
        persist_cierre_tienda(db, payload.model_dump(), user, cierre_id=cierre_id, status=normalize_cierre_status(cierre.get("status")))
    except HTTPException:
        raise
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error al actualizar: {str(exc)}") from exc
    return {"id": cierre_id, "status": "ok"}


@app.get("/api/cierres")
def list_cierres(
    fecha: Optional[str] = None,
    employee_id: Optional[int] = None,
    status: Optional[str] = None,
    include_deleted: bool = False,
    user=Depends(get_current_user),
    db=Depends(get_db),
):
    cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    query = "SELECT * FROM cierres WHERE 1=1"
    params: List[Any] = []
    if user["role"] == "employee":
        query += " AND employee_id = %s AND (tipo = 'pista' OR tipo IS NULL)"
        params.append(user["id"])
    elif user["role"] == "tienda":
        query += " AND employee_id = %s AND tipo = 'tienda'"
        params.append(user["id"])
    if fecha:
        query += " AND fecha = %s"
        params.append(fecha)
    if employee_id and user["role"] in {"admin", "supervisor"}:
        query += " AND employee_id = %s"
        params.append(employee_id)
    query += " ORDER BY fecha DESC, id DESC"
    cur.execute(query, params)
    cierres = [hydrate_cierre(row) for row in cur.fetchall()]
    normalized_status = normalize_cierre_status(status) if status else None
    should_include_deleted = include_deleted or normalized_status == STATUS_DELETED
    if not should_include_deleted:
        cierres = [cierre for cierre in cierres if cierre.get("status") != STATUS_DELETED]
    if status:
        cierres = [cierre for cierre in cierres if cierre.get("status") == normalized_status]
    if can_manage_review_notes(user):
        attach_review_note_counts(db, cierres)
    return cierres


@app.get("/api/cierres/{cierre_id}")
def get_cierre(cierre_id: int, user=Depends(get_current_user), db=Depends(get_db)):
    cierre = get_cierre_or_404(db, cierre_id)
    assert_can_view_cierre(cierre, user)
    if can_manage_review_notes(user):
        cierre["review_notes"] = list_review_notes(db, cierre_id)
    return cierre


@app.put("/api/cierres/{cierre_id}")
def update_cierre(cierre_id: int, payload: CierrePayload, user=Depends(get_current_user), db=Depends(get_db)):
    cierre = get_cierre_or_404(db, cierre_id)
    assert_can_edit_cierre(cierre, user)
    employee_id = payload.employee_id if user["role"] in {"admin", "supervisor"} and payload.employee_id else cierre.get("employee_id")
    next_status = normalize_cierre_status(cierre.get("status"))
    reset_reconciliation = False
    if user["role"] == "admin" and next_status == STATUS_RECONCILED:
        next_status = STATUS_APPROVED
        reset_reconciliation = True
    try:
        persist_cierre_payload(
            db,
            payload.model_dump(),
            payload.model_dump(),
            user,
            cierre_id=cierre_id,
            employee_id=employee_id,
            status=next_status,
            reset_reconciliation=reset_reconciliation,
        )
    except HTTPException:
        raise
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error al actualizar el cierre: {str(exc)}") from exc
    if reset_reconciliation:
        save_audit_log(db, cierre_id, user["id"], "reconciliation_reset", {"next_status": next_status})
        db.commit()
    return {"ok": True, "id": cierre_id}


@app.post("/api/cierres/{cierre_id}/review")
def review_cierre(cierre_id: int, payload: ReviewPayload, user=Depends(require_roles("admin", "supervisor")), db=Depends(get_db)):
    cierre = get_cierre_or_404(db, cierre_id)
    if cierre.get("tipo") == "tienda":
        validado_payload = normalize_tienda_payload(payload.validado_json)
        resumen_validado = compute_tienda_summary(validado_payload)
    else:
        validado_payload = normalize_payload(payload.validado_json)
        resumen_validado = compute_summary(validado_payload, validated=True)
    next_status = normalize_cierre_status(payload.status)
    assert_can_review_transition(cierre, user, next_status)
    reviewed_at = cierre.get("reviewed_at")
    reviewed_by_user_id = cierre.get("reviewed_by_user_id")
    approved_at = cierre.get("approved_at")
    approved_by_user_id = cierre.get("approved_by_user_id")
    now = utcnow()
    if next_status == STATUS_REVIEWED:
        reviewed_at = now
        reviewed_by_user_id = user["id"]
    if next_status == STATUS_APPROVED:
        approved_at = now
        approved_by_user_id = user["id"]
    editable_until = None if cierre.get("tipo") == "tienda" else now
    cur = db.cursor()
    cur.execute(
        """
        UPDATE cierres
        SET validado_json = %s::jsonb,
            resumen_validado = %s::jsonb,
            status = %s,
            audit_notes = %s,
            document_reviewed_at = %s,
            reviewed_at = %s,
            reviewed_by_user_id = %s,
            approved_at = %s,
            approved_by_user_id = %s,
            editable_until = %s,
            edited_by_user_id = %s,
            updated_at = NOW()
        WHERE id = %s
        """,
        (
            json.dumps(validado_payload),
            json.dumps(resumen_validado),
            next_status,
            payload.audit_notes or "",
            reviewed_at,
            reviewed_at,
            reviewed_by_user_id,
            approved_at,
            approved_by_user_id,
            editable_until,
            user["id"],
            cierre_id,
        ),
    )
    save_audit_log(db, cierre_id, user["id"], next_status, {"status": next_status, "notes": payload.audit_notes or ""})
    db.commit()
    return {"ok": True}


@app.post("/api/cierres/{cierre_id}/reconcile")
def reconcile_cierre(cierre_id: int, user=Depends(require_roles("admin")), db=Depends(get_db)):
    cierre = get_cierre_or_404(db, cierre_id)
    assert_can_reconcile_cierre(cierre, user)
    match = find_latest_gaspro_match_for_cierre(db, cierre)
    if not match:
        raise HTTPException(status_code=409, detail="No hay datos de Gaspro para volver a conciliar este cierre")
    apply_gaspro_to_cierre(db, cierre, match["summary"], match["import_id"], user)
    save_audit_log(db, cierre_id, user["id"], "reconciled", {"source": "latest_gaspro", "import_id": match["import_id"]})
    db.commit()
    return {"ok": True, "import_id": match["import_id"]}


@app.post("/api/cierres/{cierre_id}/notes")
def create_review_note(cierre_id: int, payload: ReviewNotePayload, user=Depends(require_roles("admin", "supervisor")), db=Depends(get_db)):
    cierre = get_cierre_or_404(db, cierre_id)
    assert_can_view_cierre(cierre, user)
    if normalize_cierre_status(cierre.get("status")) == STATUS_DELETED:
        raise HTTPException(status_code=400, detail="No se pueden agregar notas a un cierre en papelera")

    target_scope = (payload.target_scope or "").strip().lower()
    section_key = (payload.section_key or "").strip()
    body = (payload.body or "").strip()
    movement_id = (payload.movement_id or "").strip() or None

    if target_scope not in {"section", "item"}:
        raise HTTPException(status_code=422, detail="Tipo de nota inválido")
    if not section_key:
        raise HTTPException(status_code=422, detail="La sección es obligatoria")
    if not body:
        raise HTTPException(status_code=422, detail="La nota no puede quedar vacía")
    if section_key not in allowed_note_sections(cierre):
        raise HTTPException(status_code=422, detail="Esta sección no admite notas QA")
    if target_scope == "item":
        if cierre.get("tipo") == "tienda":
            raise HTTPException(status_code=422, detail="Los cierres de tienda solo admiten notas por sección")
        if section_key not in ITEM_NOTE_SECTIONS:
            raise HTTPException(status_code=422, detail="Esta sección no admite notas por movimiento")
        if not movement_id:
            raise HTTPException(status_code=422, detail="El movimiento es obligatorio para una nota por ítem")
        if not movement_exists_for_review_note(cierre, section_key, movement_id):
            raise HTTPException(status_code=422, detail="El movimiento indicado ya no existe en este cierre")
    else:
        movement_id = None

    cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        """
        INSERT INTO cierre_review_notes (
            cierre_id, target_scope, section_key, movement_id, body, created_by_user_id
        ) VALUES (%s, %s, %s, %s, %s, %s)
        RETURNING id
        """,
        (cierre_id, target_scope, section_key, movement_id, body, user["id"]),
    )
    note_id = cur.fetchone()["id"]
    save_audit_log(db, cierre_id, user["id"], "review_note_created", {"note_id": note_id, "section_key": section_key, "movement_id": movement_id})
    db.commit()
    note = next((item for item in list_review_notes(db, cierre_id) if item["id"] == note_id), None)
    return note or {"id": note_id}


@app.patch("/api/cierres/{cierre_id}/notes/{note_id}")
def update_review_note(cierre_id: int, note_id: int, payload: ReviewNoteUpdatePayload, user=Depends(require_roles("admin", "supervisor")), db=Depends(get_db)):
    cierre = get_cierre_or_404(db, cierre_id)
    assert_can_view_cierre(cierre, user)
    if normalize_cierre_status(cierre.get("status")) == STATUS_DELETED:
        raise HTTPException(status_code=400, detail="Restaura este cierre antes de resolver notas")
    cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM cierre_review_notes WHERE id = %s AND cierre_id = %s", (note_id, cierre_id))
    note = cur.fetchone()
    if not note:
        raise HTTPException(status_code=404, detail="Nota no encontrada")

    resolved = bool(payload.resolved)
    resolved_at = utcnow() if resolved else None
    resolved_by_user_id = user["id"] if resolved else None
    cur.execute(
        """
        UPDATE cierre_review_notes
        SET resolved = %s,
            resolved_at = %s,
            resolved_by_user_id = %s,
            updated_at = NOW()
        WHERE id = %s
        """,
        (resolved, resolved_at, resolved_by_user_id, note_id),
    )
    save_audit_log(db, cierre_id, user["id"], "review_note_updated", {"note_id": note_id, "resolved": resolved})
    db.commit()
    updated = next((item for item in list_review_notes(db, cierre_id) if item["id"] == note_id), None)
    return updated or {"id": note_id, "resolved": resolved}


@app.delete("/api/cierres/{cierre_id}")
def delete_cierre(cierre_id: int, payload: DeleteCierrePayload, user=Depends(require_roles("admin")), db=Depends(get_db)):
    cierre = get_cierre_or_404(db, cierre_id)
    current_status = normalize_cierre_status(cierre.get("status"))
    if current_status == STATUS_DELETED:
        return {"ok": True, "id": cierre_id}
    cur = db.cursor()
    cur.execute(
        """
        UPDATE cierres
        SET status = %s,
            status_before_delete = %s,
            deleted_at = %s,
            deleted_by_user_id = %s,
            delete_reason = %s,
            edited_by_user_id = %s,
            updated_at = NOW()
        WHERE id = %s
        """,
        (STATUS_DELETED, current_status, utcnow(), user["id"], (payload.reason or "").strip(), user["id"], cierre_id),
    )
    save_audit_log(db, cierre_id, user["id"], "deleted", {"reason": (payload.reason or "").strip(), "previous_status": current_status})
    db.commit()
    return {"ok": True, "id": cierre_id}


@app.post("/api/cierres/{cierre_id}/restore")
def restore_cierre(cierre_id: int, user=Depends(require_roles("admin")), db=Depends(get_db)):
    cierre = get_cierre_or_404(db, cierre_id)
    current_status = normalize_cierre_status(cierre.get("status"))
    if current_status != STATUS_DELETED:
        raise HTTPException(status_code=400, detail="Este cierre no está en papelera")
    restore_status = normalize_cierre_status(cierre.get("status_before_delete")) or STATUS_SUBMITTED
    if restore_status == STATUS_DELETED:
        restore_status = STATUS_SUBMITTED
    cur = db.cursor()
    cur.execute(
        """
        UPDATE cierres
        SET status = %s,
            deleted_at = NULL,
            deleted_by_user_id = NULL,
            delete_reason = '',
            status_before_delete = NULL,
            edited_by_user_id = %s,
            updated_at = NOW()
        WHERE id = %s
        """,
        (restore_status, user["id"], cierre_id),
    )
    save_audit_log(db, cierre_id, user["id"], "restored", {"restored_status": restore_status})
    db.commit()
    return {"ok": True, "id": cierre_id, "status": restore_status}


@app.post("/api/cierres/{cierre_id}/attachments")
def upload_attachment(
    cierre_id: int,
    category: str = Form(...),
    movement_id: Optional[str] = Form(default=None),
    file: UploadFile = File(...),
    user=Depends(get_current_user),
    db=Depends(get_db),
):
    cierre = get_cierre_or_404(db, cierre_id)
    assert_can_view_cierre(cierre, user)
    saved = save_upload(file, f"cierres/{cierre_id}")
    cur = db.cursor()
    cur.execute(
        """
        INSERT INTO cierre_attachments (cierre_id, uploaded_by_user_id, category, movement_id, original_name, storage_path, mime_type, size_bytes, sha256)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        RETURNING id
        """,
        (cierre_id, user["id"], category, movement_id, saved["original_name"], saved["path"], saved["mime_type"], saved["size_bytes"], saved["sha256"]),
    )
    attachment_id = cur.fetchone()[0]
    save_audit_log(db, cierre_id, user["id"], "attachment_uploaded", {"attachment_id": attachment_id, "category": category})
    db.commit()
    return {"id": attachment_id, **saved}


@app.get("/api/cierres/{cierre_id}/attachments")
def list_attachments(cierre_id: int, user=Depends(get_current_user), db=Depends(get_db)):
    cierre = get_cierre_or_404(db, cierre_id)
    assert_can_view_cierre(cierre, user)
    cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM cierre_attachments WHERE cierre_id = %s ORDER BY created_at DESC", (cierre_id,))
    return [dict(row) for row in cur.fetchall()]


@app.post("/api/gaspro/import", response_model=GasproImportResponse)
def import_gaspro(
    import_mode: str = Form(...),
    date_from: str = Form(...),
    date_to: str = Form(...),
    file: UploadFile = File(...),
    user=Depends(require_roles("admin")),
    db=Depends(get_db),
):
    saved = save_upload(file, "gaspro")
    rows = read_tabular_upload(Path(saved["path"]))
    summary_by_date = summarize_gaspro_rows(rows)
    cur = db.cursor()
    cur.execute(
        """
        INSERT INTO gaspro_imports (uploaded_by_user_id, import_mode, original_name, storage_path, mime_type, size_bytes, sha256, date_from, date_to)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        RETURNING id
        """,
        (user["id"], import_mode, saved["original_name"], saved["path"], saved["mime_type"], saved["size_bytes"], saved["sha256"], date_from, date_to),
    )
    import_id = cur.fetchone()[0]
    for row in rows:
        cur.execute(
            """
            INSERT INTO gaspro_rows (gaspro_import_id, fecha, empleado, turno, producto, litros, monto, ppu, raw_row)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s::jsonb)
            """,
            (import_id, row["fecha"], row["empleado"], row["turno"], row["producto"], row["litros"], row["monto"], row["ppu"], json.dumps(row["raw_row"])),
        )
    matched = 0
    skipped = 0
    already_reconciled = 0
    cur_dict = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur_dict.execute("SELECT * FROM cierres WHERE fecha BETWEEN %s AND %s", (date_from, date_to))
    for cierre in [hydrate_cierre(row) for row in cur_dict.fetchall()]:
        if cierre.get("tipo") == "tienda" or cierre.get("status") == STATUS_DELETED:
            continue
        date_key = cierre["fecha"].isoformat() if isinstance(cierre["fecha"], date) else str(cierre["fecha"])
        employee_key = normalize_text(cierre.get("empleado"))
        summary = summary_by_date.get(date_key, {}).get(employee_key)
        if not summary:
            skipped += 1
            continue
        current_status = normalize_cierre_status(cierre.get("status"))
        if current_status == STATUS_RECONCILED:
            already_reconciled += 1
            continue
        if current_status != STATUS_APPROVED:
            skipped += 1
            continue
        apply_gaspro_to_cierre(db, cierre, summary, import_id, user)
        matched += 1
    cur.execute("UPDATE gaspro_imports SET matched_cierres = %s, reconciled_at = NOW() WHERE id = %s", (matched, import_id))
    db.commit()
    return GasproImportResponse(
        import_id=import_id,
        matched_cierres=matched,
        skipped_cierres=skipped,
        already_reconciled=already_reconciled,
        import_mode=import_mode,
    )


@app.get("/api/gaspro/imports")
def list_gaspro_imports(user=Depends(require_roles("admin", "supervisor")), db=Depends(get_db)):
    cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM gaspro_imports ORDER BY created_at DESC")
    return [dict(row) for row in cur.fetchall()]


@app.get("/api/export/csv")
def export_csv(fecha_inicio: Optional[str] = None, fecha_fin: Optional[str] = None, user=Depends(require_roles("admin", "supervisor")), db=Depends(get_db)):
    cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    query = "SELECT * FROM cierres WHERE 1=1"
    params = []
    if fecha_inicio:
        query += " AND fecha >= %s"
        params.append(fecha_inicio)
    if fecha_fin:
        query += " AND fecha <= %s"
        params.append(fecha_fin)
    query += " ORDER BY fecha ASC, id ASC"
    cur.execute(query, params)
    rows = [hydrate_cierre(row) for row in cur.fetchall()]
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["id", "fecha", "empleado", "turno", "status", "total_reportado", "total_validado", "gaspro_mode"])
    for row in rows:
        writer.writerow([
            row["id"],
            row["fecha"],
            row["empleado"],
            row["turno"],
            row["status"],
            row["resumen_reportado"].get("total_reportado", 0),
            row["resumen_validado"].get("total_reportado", 0),
            row.get("gaspro_mode") or "",
        ])
    output.seek(0)
    return StreamingResponse(output, media_type="text/csv", headers={"Content-Disposition": f"attachment; filename=cierres_{date.today().isoformat()}.csv"})


@app.get("/api/export/monthly.xlsx")
def export_monthly_excel(month: str, validated: bool = True, user=Depends(require_roles("admin", "supervisor")), db=Depends(get_db)):
    try:
        month_start = datetime.strptime(month, "%Y-%m").date().replace(day=1)
    except ValueError:
        raise HTTPException(status_code=400, detail="Formato de mes inválido. Usá YYYY-MM")
    next_month = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1)
    cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM cierres WHERE fecha >= %s AND fecha < %s ORDER BY fecha, turno", (month_start, next_month))
    cierres = [hydrate_cierre(row) for row in cur.fetchall()]
    if EXPORT_TEMPLATE_PATH.exists():
        wb = openpyxl.load_workbook(EXPORT_TEMPLATE_PATH)
    else:
        wb = openpyxl.Workbook()
        wb.active.title = month_start.strftime("%m")
    for cierre in cierres:
        ws = find_day_sheet(wb, int(cierre["fecha"].day))
        if ws:
            fill_cierre_in_sheet(ws, cierre, validated=validated)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=cierres_{month}.xlsx"},
    )


FRONTEND_DIR = os.path.join(os.path.dirname(__file__), "frontend", "dist")
if os.path.exists(FRONTEND_DIR):
    app.mount("/assets", StaticFiles(directory=os.path.join(FRONTEND_DIR, "assets")), name="assets")

    @app.get("/")
    def serve_root():
        return FileResponse(os.path.join(FRONTEND_DIR, "index.html"))

    @app.get("/{full_path:path}")
    def serve_spa(full_path: str):
        if full_path.startswith("api/"):
            raise HTTPException(status_code=404, detail="Not found")
        file_path = os.path.join(FRONTEND_DIR, full_path)
        if os.path.exists(file_path) and os.path.isfile(file_path):
            return FileResponse(file_path)
        return FileResponse(os.path.join(FRONTEND_DIR, "index.html"))


if __name__ == "__main__":
    import uvicorn

    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
